import json
import re
from openai import OpenAI
import pandas as pd
from pathlib import Path
from django.conf import settings
from PyPDF2 import PdfReader
from docx import Document
import os
import requests
from fuzzywuzzy import fuzz
from collections import defaultdict, Counter
import numpy as np
from io import BytesIO
import base64
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import logging
from urllib.parse import urlencode

logger = logging.getLogger(__name__)

# ============================================================================
# CACHES AND CONSTANTS
# ============================================================================

DESIGNATION_CACHE = {}
SKILL_RELEVANCE_CACHE = {}

# Global token usage tracker (resets on server restart)
GLOBAL_TOKEN_USAGE = {
    'total_tokens': 0,
    'total_cost': 0.0,
    'operations': defaultdict(lambda: {'tokens': 0, 'calls': 0, 'cost': 0.0})
}

def track_global_tokens(operation, tokens, model='gpt-4o-mini'):
    """Track tokens globally across all operations"""
    pricing = {
        'gpt-4o-mini': {'prompt': 0.00015, 'completion': 0.0006},  # per 1K tokens
        'gpt-4': {'prompt': 0.03, 'completion': 0.06},
    }
    
    model_pricing = pricing.get(model, pricing['gpt-4o-mini'])
    cost = (tokens.get('prompt_tokens', 0) / 1000 * model_pricing['prompt'] + 
            tokens.get('completion_tokens', 0) / 1000 * model_pricing['completion'])
    
    GLOBAL_TOKEN_USAGE['total_tokens'] += tokens.get('total_tokens', 0)
    GLOBAL_TOKEN_USAGE['total_cost'] += cost
    GLOBAL_TOKEN_USAGE['operations'][operation]['tokens'] += tokens.get('total_tokens', 0)
    GLOBAL_TOKEN_USAGE['operations'][operation]['calls'] += 1
    GLOBAL_TOKEN_USAGE['operations'][operation]['cost'] += cost


GENERIC_SOFT_SKILLS = {
    "communication", "leadership", "teamwork", "problem solving", 
    "collaboration", "presentation", "people management",
    "stakeholder management"
}

# ============================================================================
# CORE UTILITY FUNCTIONS
# ============================================================================

def parse_experience_years(exp_str):
    '''Extract numeric years from experience string'''
    if not exp_str or str(exp_str).lower() in ['nan', 'none', 'n/a', '']:
        return None
    
    exp_str = str(exp_str).lower()
    
    # Pattern 1: "X years" or "X+ years"
    match = re.search(r'(\d+\.?\d*)\+?\s*(?:years?|yrs?)', exp_str)
    if match:
        return float(match.group(1))
    
    # Pattern 2: "X-Y years" (take average)
    match = re.search(r'(\d+\.?\d*)\s*-\s*(\d+\.?\d*)\s*(?:years?|yrs?)', exp_str)
    if match:
        return (float(match.group(1)) + float(match.group(2))) / 2
    
    # Pattern 3: Just a number
    match = re.search(r'(\d+\.?\d*)', exp_str)
    if match:
        return float(match.group(1))
    
    return None

def generate_role_keyword_profile(jd_role_title: str):
    """Extract domain-relevant keywords from role title."""
    if not jd_role_title:
        return []

    jd = jd_role_title.lower()
    remove_words = {
        "senior", "jr", "junior", "lead", "manager", "associate", "intern",
        "executive", "specialist", "head", "chief", "vp", "director"
    }

    words = re.split(r"[ /,|-]+", jd)
    keywords = [w.strip() for w in words if len(w) > 2 and w not in remove_words]

    return keywords

# ============================================================================
# AI-POWERED ANALYSIS FUNCTIONS - OPTIMIZED WITH BATCH PROCESSING
# ============================================================================

def call_openai_analysis(prompt, system_message, temperature=0.2, max_tokens=300):
    """Unified OpenAI API call with error handling and token tracking"""
    try:
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": prompt}
            ],
            temperature=temperature,
            max_tokens=max_tokens
        )
        
        content = response.choices[0].message.content.strip()
        cleaned = re.sub(r"^```json\s*|\s*```$", "", content, flags=re.MULTILINE).strip()
        
        # Extract token usage
        token_usage = {
            'prompt_tokens': response.usage.prompt_tokens,
            'completion_tokens': response.usage.completion_tokens,
            'total_tokens': response.usage.total_tokens
        }
        
        result = json.loads(cleaned)
        
        # Add token usage to result
        if isinstance(result, dict):
            result['_token_usage'] = token_usage
        
        return result
    
    except json.JSONDecodeError as je:
        print(f"⚠️ JSON Decode Error: {je}")
        print(f"⚠️ Raw Output: {content}")
        return {}
    except Exception as e:
        print(f"⚠️ OpenAI API error: {e}")
        return {}

# ============================================================================
# BATCH PROCESSING FUNCTIONS - NEW!
# ============================================================================

def batch_analyze_designations(jd_role, candidate_designations, batch_size=20):
    """
    Batch process multiple candidate designations in a SINGLE API call.
    This reduces 200 API calls to just 10 calls (for 200 candidates).
    
    Returns: dict mapping candidate_designation -> (score, match_type, details)
    """
    if not jd_role or not candidate_designations:
        return {}
    
    jd_role = str(jd_role).lower().strip()
    results = {}
    
    # Filter out already cached results
    uncached_designations = []
    for cand_desg in candidate_designations:
        cand_desg_clean = str(cand_desg).lower().strip()
        
        # Skip invalid
        if not cand_desg_clean or cand_desg_clean in ['nan', 'none', 'n/a', '']:
            results[cand_desg] = (0, "no_data", {})
            continue
        
        # Check exact match
        if jd_role == cand_desg_clean:
            results[cand_desg] = (100, "exact", {"matched": "exact designation match", "cache_hit": False})
            continue
        
        # Check cache
        cache_key = f"{jd_role}||{cand_desg_clean}"
        if cache_key in DESIGNATION_CACHE:
            cached_result = DESIGNATION_CACHE[cache_key]
            if len(cached_result) >= 3 and isinstance(cached_result[2], dict):
                cached_result[2]['cache_hit'] = True
            results[cand_desg] = cached_result
        else:
            uncached_designations.append(cand_desg_clean)
    
    if not uncached_designations:
        print(f"✅ All {len(candidate_designations)} designations found in cache!")
        return results
    
    print(f"🔄 Batch analyzing {len(uncached_designations)} designations (cached: {len(results)})...")
    
    # Process in batches
    for i in range(0, len(uncached_designations), batch_size):
        batch = uncached_designations[i:i + batch_size]
        
        # Create batch prompt
        designations_list = "\n".join([f'{idx+1}. "{desg}"' for idx, desg in enumerate(batch)])
        
        prompt = f'''You are an expert HR analyst. Compare the following candidate designations against the target job role.

TARGET JOB ROLE: "{jd_role}"

CANDIDATE DESIGNATIONS:
{designations_list}

For EACH designation, analyze similarity and return a JSON array with this structure:

[
  {{
    "designation": "<exact designation from list>",
    "similarity_score": <0-100>,
    "match_type": "<exact|high|moderate|low|no_match>",
    "confidence": "<high|medium|low>",
    "reasoning": "<brief 1 sentence>",
    "seniority_match": <true|false>,
    "function_match": <true|false>,
    "role_equivalent": <true|false>
  }},
  ...
]

Return ONLY valid JSON array (no markdown, no code blocks).'''

        result = call_openai_analysis(
            prompt,
            "You are an expert HR analyst. Analyze job role similarities efficiently. Return only valid JSON.",
            temperature=0.1,
            max_tokens=batch_size * 100  # Scale tokens with batch size
        )
        
        # Track tokens for this batch
        if isinstance(result, dict) and '_token_usage' in result:
            track_global_tokens('batch_designation_analysis', result['_token_usage'])
        
        # Parse results
        if isinstance(result, list):
            batch_results = result
        elif isinstance(result, dict) and 'results' in result:
            batch_results = result['results']
        else:
            print(f"⚠️ Unexpected batch result format, using fallback")
            batch_results = []
        
        # Map results back
        for item in batch_results:
            if not isinstance(item, dict):
                continue
                
            designation = item.get('designation', '').lower().strip()
            score = float(item.get('similarity_score', 0))
            score = max(0, min(100, score))
            match_type = item.get('match_type', 'unknown')
            
            details = {
                'reasoning': item.get('reasoning', 'No reasoning provided'),
                'confidence': item.get('confidence', 'unknown'),
                'seniority_match': item.get('seniority_match', False),
                'function_match': item.get('function_match', False),
                'role_equivalent': item.get('role_equivalent', False),
                'api_source': 'openai_batch',
                'cache_hit': False,
                'batch_processed': True
            }
            
            result_tuple = (score, match_type, details)
            
            # Store in cache
            cache_key = f"{jd_role}||{designation}"
            DESIGNATION_CACHE[cache_key] = result_tuple
            
            # Store in results (match original designation)
            for orig_desg in candidate_designations:
                if str(orig_desg).lower().strip() == designation:
                    results[orig_desg] = result_tuple
                    break
    
    return results


def batch_analyze_skill_relevance(jd_role_title, candidates_with_skills, batch_size=15):
    """
    Batch process skill relevance for multiple candidates in SINGLE API calls.
    
    Args:
        jd_role_title: Job role title
        candidates_with_skills: List of tuples [(candidate_designation, [matched_skills]), ...]
        batch_size: Number of candidates per batch
    
    Returns: dict mapping (designation, skills_key) -> relevance_score
    """
    if not jd_role_title or not candidates_with_skills:
        return {}
    
    results = {}
    uncached_items = []
    
    # Check cache first
    for designation, matched_skills in candidates_with_skills:
        if not matched_skills:
            results[(designation, tuple(matched_skills))] = 0
            continue
        
        skills_key = ",".join(sorted(matched_skills[:10]))
        cache_key = f"{jd_role_title}||{skills_key}"
        
        if cache_key in SKILL_RELEVANCE_CACHE:
            cached = SKILL_RELEVANCE_CACHE[cache_key]
            if isinstance(cached, dict) and 'score' in cached:
                results[(designation, tuple(matched_skills))] = cached['score']
            else:
                results[(designation, tuple(matched_skills))] = cached
        else:
            uncached_items.append((designation, matched_skills, cache_key))
    
    if not uncached_items:
        print(f"✅ All {len(candidates_with_skills)} skill relevance scores found in cache!")
        return results
    
    print(f"🔄 Batch analyzing skill relevance for {len(uncached_items)} candidates (cached: {len(results)})...")
    
    # Process in batches
    for i in range(0, len(uncached_items), batch_size):
        batch = uncached_items[i:i + batch_size]
        
        # Create batch prompt
        candidates_list = []
        for idx, (designation, skills, _) in enumerate(batch):
            skills_str = ", ".join(skills[:15])
            candidates_list.append(f'{idx+1}. Designation: "{designation}", Skills: {skills_str}')
        
        candidates_text = "\n".join(candidates_list)
        
        prompt = f'''You are an expert technical recruiter. Analyze skill-to-role fit for multiple candidates.

JOB ROLE: "{jd_role_title}"

CANDIDATES:
{candidates_text}

For EACH candidate, analyze how relevant their matched skills are for this job role.

SCORING (0-100):
- 80-100: Excellent - strong core skills directly relevant
- 60-79: Good - relevant skills with domain match
- 45-59: Moderate - some relevance, decent transferability
- 30-44: Fair - limited relevance, mostly generic
- 0-29: Poor - wrong domain or no relevant skills

Return JSON array:
[
  {{
    "candidate_number": <1-{len(batch)}>,
    "relevance_score": <0-100>,
    "quality": "<high|medium|low>",
    "reasoning": "<brief 1-2 sentences>",
    "core_skills_present": <true|false>,
    "domain_match": <true|false>
  }},
  ...
]

Return ONLY valid JSON array (no markdown, no code blocks).'''

        result = call_openai_analysis(
            prompt,
            "You are an expert recruiter. Analyze skill relevance efficiently. Return only valid JSON.",
            temperature=0.2,
            max_tokens=batch_size * 150
        )
        
        # Track tokens
        if isinstance(result, dict) and '_token_usage' in result:
            track_global_tokens('batch_skill_relevance', result['_token_usage'])
        
        # Parse results
        if isinstance(result, list):
            batch_results = result
        elif isinstance(result, dict) and 'results' in result:
            batch_results = result['results']
        else:
            batch_results = []
        
        # Map results back
        for item in batch_results:
            if not isinstance(item, dict):
                continue
            
            cand_num = item.get('candidate_number', 0) - 1
            if cand_num < 0 or cand_num >= len(batch):
                continue
            
            designation, skills, cache_key = batch[cand_num]
            
            relevance_score = float(item.get('relevance_score', 50))
            relevance_score = max(0, min(100, relevance_score))
            
            # Cache the result
            cache_value = {
                'score': relevance_score,
                'details': item,
                'batch_processed': True
            }
            SKILL_RELEVANCE_CACHE[cache_key] = cache_value
            
            results[(designation, tuple(skills))] = relevance_score
    
    return results


# ============================================================================
# MODIFIED FUNCTIONS - NOW USE BATCH PROCESSING
# ============================================================================

def is_designation_relevant(jd_role, candidate_designation, threshold=30):
    '''Use OpenAI to check if candidate designation is relevant to JD role - BALANCED VERSION'''
    if not jd_role or not candidate_designation:
        return True
    
    jd_role = str(jd_role).lower().strip()
    cand_desg = str(candidate_designation).lower().strip()
    
    if cand_desg in ['nan', 'none', 'n/a', '']:
        return True
    
    # Quick exact/substring match
    if jd_role == cand_desg or jd_role in cand_desg or cand_desg in jd_role:
        return True
    
    # Check cache first
    cache_key = f"{jd_role}||{cand_desg}"
    if cache_key in DESIGNATION_CACHE:
        cached_result = DESIGNATION_CACHE[cache_key]
        score = cached_result[0] if len(cached_result) > 0 else 0
        return score >= threshold
    
    # If not in cache, we'll batch process later
    # For now, be lenient and return True
    return True


def calculate_skill_relevance_score(matched_skills, jd_role_title, candidate_designation, use_cache=True):
    '''Use OpenAI to calculate how relevant the matched skills are to the role - OPTIMIZED for batch processing'''
    if not matched_skills:
        return 0
    
    if not jd_role_title:
        return 50
    
    # Create cache key
    skills_key = ",".join(sorted(matched_skills[:10]))
    cache_key = f"{jd_role_title}||{skills_key}"
    
    if use_cache and cache_key in SKILL_RELEVANCE_CACHE:
        cached_result = SKILL_RELEVANCE_CACHE[cache_key]
        if isinstance(cached_result, dict) and 'score' in cached_result:
            return cached_result['score']
        return cached_result
    
    # If not in cache, return default score
    # Batch processing will fill this later
    return 50


def calculate_designation_similarity(jd_role, candidate_designation, use_fuzzy=True, use_cache=True):
    '''Calculate similarity between JD role and candidate designation - OPTIMIZED for batch processing'''
    if not jd_role or not candidate_designation:
        return 0, "no_data", {}
    
    jd_role = str(jd_role).lower().strip()
    cand_desg = str(candidate_designation).lower().strip()
    
    if not jd_role or not cand_desg or cand_desg in ['nan', 'none', 'n/a', '']:
        return 0, "no_data", {}
    
    # Quick exact match check
    if jd_role == cand_desg:
        return 100, "exact", {"matched": "exact designation match", "cache_hit": False}
    
    # Check cache
    cache_key = f"{jd_role}||{cand_desg}"
    if use_cache and cache_key in DESIGNATION_CACHE:
        cached_result = DESIGNATION_CACHE[cache_key]
        if len(cached_result) >= 3 and isinstance(cached_result[2], dict):
            cached_result[2]['cache_hit'] = True
        return cached_result
    
    # Fallback to fuzzy matching if not in cache
    return fallback_designation_matching(jd_role, cand_desg, use_cache, cache_key)


def fallback_designation_matching(jd_role, cand_desg, use_cache, cache_key):
    '''Fallback designation matching using fuzzy logic'''
    
    # Substring match
    if jd_role in cand_desg or cand_desg in jd_role:
        result = (90, "substring", {"matched": "substring match (fallback)", "cache_hit": False})
        if use_cache:
            DESIGNATION_CACHE[cache_key] = result
        return result
    
    # Fuzzy string matching
    token_sort = fuzz.token_sort_ratio(jd_role, cand_desg)
    partial = fuzz.partial_ratio(jd_role, cand_desg)
    token_set = fuzz.token_set_ratio(jd_role, cand_desg)
    
    best_fuzzy = max(token_sort, partial, token_set)
    
    if best_fuzzy >= 80:
        score = 70 + ((best_fuzzy - 80) * 1.5)
        result = (score, "fuzzy_high", {"similarity": best_fuzzy, "source": "fallback", "cache_hit": False})
    elif best_fuzzy >= 70:
        score = 50 + ((best_fuzzy - 70) * 2)
        result = (score, "fuzzy_medium", {"similarity": best_fuzzy, "source": "fallback", "cache_hit": False})
    elif best_fuzzy >= 60:
        score = 30 + ((best_fuzzy - 60) * 2)
        result = (score, "fuzzy_low", {"similarity": best_fuzzy, "source": "fallback", "cache_hit": False})
    else:
        # Word overlap
        jd_words = set(jd_role.split())
        cand_words = set(cand_desg.split())
        
        stopwords = {'and', 'or', 'the', 'a', 'an', 'of', 'in', 'to', 'for', 'with', 'on', 'at', 'by'}
        jd_words_clean = jd_words - stopwords
        cand_words_clean = cand_words - stopwords
        
        if jd_words_clean and cand_words_clean:
            common = jd_words_clean & cand_words_clean
            if common:
                union = jd_words_clean | cand_words_clean
                jaccard = len(common) / len(union)
                score = min(40, jaccard * 60)
                result = (score, "word_overlap", {
                    "common_words": list(common),
                    "jaccard": jaccard,
                    "source": "fallback",
                    "cache_hit": False
                })
            else:
                result = (0, "no_match", {"source": "fallback", "cache_hit": False})
        else:
            result = (0, "no_match", {"source": "fallback", "cache_hit": False})
    
    if use_cache:
        DESIGNATION_CACHE[cache_key] = result
    return result

# ============================================================================
# SKILL MATCHING ENGINE
# ============================================================================
def calculate_skill_similarity(req_skill, cand_skill, use_fuzzy=True):
    '''Direct exact matching only - 100 or 0'''
    req_norm = normalize_skill(req_skill)
    cand_norm = normalize_skill(cand_skill)
    
    # Exact match only
    if req_norm == cand_norm:
        return 100, "exact", {"matched": "exact match"}
    
    # Optional: Substring match for compound skills
    if req_norm in cand_norm or cand_norm in req_norm:
        return 85, "substring", {"matched": "partial match"}
    
    return 0, "no_match", {}

def normalize_skill(skill):
    '''
    Enhanced normalization for better matching
    - Lowercase
    - Remove special characters (/,-,etc.)
    - Collapse multiple spaces
    - Strip whitespace
    '''
    if not skill:
        return ""
    
    skill = str(skill).lower().strip()
    
    # Remove special characters but keep spaces
    skill = re.sub(r'[^\w\s]', ' ', skill)
    
    # Collapse multiple spaces
    skill = ' '.join(skill.split())
    
    return skill

def skill_tokens(skill):
    '''Extract meaningful word tokens from skill (2+ chars)'''
    normalized = normalize_skill(skill)
    return {word for word in normalized.split() if len(word) >= 2}

def extract_skill_versions(skill):
    '''Extract version numbers from skills'''
    version_pattern = r'(\d+(?:\.\d+)*(?:\.[xX])?)'
    match = re.search(version_pattern, skill)
    
    if match:
        base_skill = re.sub(version_pattern, '', skill).strip()
        version = match.group(1)
        return base_skill, version
    
    return skill, None

def calculate_skill_similarity(req_skill, cand_skill, use_fuzzy=True):
    '''
    Smart skill matching with multiple strategies:
    1. Exact match (100)
    2. Substring match (85)
    3. Token overlap match (70)
    4. No match (0)
    '''
    req_norm = normalize_skill(req_skill)
    cand_norm = normalize_skill(cand_skill)
    
    if not req_norm or not cand_norm:
        return 0, "no_match", {}
    
    # Strategy 1: Exact match
    if req_norm == cand_norm:
        return 100, "exact", {"matched": "exact match"}
    
    # Strategy 2: Substring match (one contains the other)
    if req_norm in cand_norm or cand_norm in req_norm:
        return 85, "substring", {"matched": "substring match"}
    
    # Strategy 3: Token overlap match
    # (e.g., "ui ux design" matches "ui design" or "ux")
    req_tokens = skill_tokens(req_skill)
    cand_tokens = skill_tokens(cand_skill)
    
    if req_tokens and cand_tokens:
        common_tokens = req_tokens & cand_tokens
        
        if common_tokens:
            # Calculate overlap percentage
            overlap_ratio = len(common_tokens) / min(len(req_tokens), len(cand_tokens))
            
            if overlap_ratio >= 0.8:  # 80%+ overlap
                return 75, "high_token_overlap", {
                    "common_tokens": list(common_tokens),
                    "overlap": f"{overlap_ratio:.0%}"
                }
            elif overlap_ratio >= 0.5:  # 50%+ overlap
                return 65, "medium_token_overlap", {
                    "common_tokens": list(common_tokens),
                    "overlap": f"{overlap_ratio:.0%}"
                }
            elif len(common_tokens) >= 1 and len(list(common_tokens)[0]) >= 4:
                # At least one meaningful word (4+ chars) matches
                return 55, "partial_token_match", {
                    "common_tokens": list(common_tokens)
                }
    
    return 0, "no_match", {}

def calculate_experience_score(candidate_exp, required_exp_range=None):
    '''Calculate experience match score (0-15 bonus points)'''
    if not required_exp_range:
        return 0
    
    cand_years = parse_experience_years(candidate_exp)
    if cand_years is None:
        return 0
    
    # Handle single number or range
    if isinstance(required_exp_range, (int, float)):
        min_years, max_years = required_exp_range, required_exp_range + 3
    elif isinstance(required_exp_range, (list, tuple)) and len(required_exp_range) == 2:
        min_years, max_years = required_exp_range
    else:
        return 0
    
    # Scoring logic
    if min_years <= cand_years <= max_years:
        return 15
    elif cand_years > max_years and cand_years <= max_years + 2:
        return 12
    elif cand_years < min_years and cand_years >= min_years - 1:
        return 10
    elif cand_years > max_years + 2 and cand_years <= max_years + 5:
        return 8
    else:
        return 0

# ============================================================================
# MODIFIED API FETCH FUNCTIONS - DYNAMIC QUERY PARAMETERS
# ============================================================================

def build_api_url(skills=None, experience=None, page=1, limit=100):
    """
    Build dynamic API URL with query parameters for Kunji API
    
    Args:
        skills: List of skills or comma-separated string
        experience: Experience in years (integer or string)
        page: Page number for pagination
        limit: Number of results per page
    
    Returns:
        Complete API URL with query parameters
    """
    base_url = getattr(settings, 'CANDIDATES_API_BASE_URL', None)
    
    if not base_url:
        raise ValueError("CANDIDATES_API_BASE_URL not configured in settings")
    
    # Build query parameters
    params = {}
    
    # Add skills parameter - Kunji API uses comma-separated skills
    if skills:
        if isinstance(skills, list):
            # Join skills with comma
            skills_str = ','.join(str(s).strip() for s in skills if s)
        else:
            skills_str = str(skills)
        
        if skills_str:
            params['skills'] = skills_str
    
    # Add experience parameter
    if experience is not None:
        # Extract numeric experience if it's a string
        if isinstance(experience, str):
            exp_years = parse_experience_years(experience)
            if exp_years is not None:
                params['experience'] = int(exp_years)
        elif isinstance(experience, (int, float)):
            params['experience'] = int(experience)
        elif isinstance(experience, dict) and 'min' in experience:
            # Use minimum experience from range
            params['experience'] = int(experience['min'])
    
    # Add pagination parameters
    params['page'] = page
    params['limit'] = limit
    
    # Build complete URL
    url = f"{base_url}?{urlencode(params)}"
    
    print(f"🔗 Built Kunji API URL: {url}")
    return url

def fetch_all_candidates_from_api(
    skills=None,
    experience=None,
    page_size=100,
    max_pages=None,
    timeout=30
):
    """
    Fetch ALL candidates from Kunji API with pagination
    FIXED: Properly handle columnar format (cols + data)
    """
    all_candidates = []
    current_page = 1
    
    print(f"🚀 Starting to fetch all candidates from Kunji API...")
    print(f"   Skills: {skills}")
    print(f"   Experience: {experience}")
    print(f"   Page size: {page_size}")
    
    while True:
        if max_pages and current_page > max_pages:
            print(f"⚠️ Reached maximum page limit ({max_pages})")
            break
        
        try:
            url = build_api_url(
                skills=skills,
                experience=experience,
                page=current_page,
                limit=page_size
            )
            
            print(f"📄 Fetching page {current_page}...")
            api_token = getattr(settings, 'CANDIDATES_API_TOKEN', None)
            
            headers = {'Content-Type': 'application/json'}
            if api_token:
                headers['Authorization'] = f'Bearer {api_token}'
            
            response = requests.get(url, headers=headers, timeout=timeout)
            response.raise_for_status()
            
            data = response.json()
            
            # Debug: Print API response structure
            if current_page == 1:
                print(f"📦 API Response keys: {list(data.keys()) if isinstance(data, dict) else 'List response'}")
            
            # ============================================================
            # FIXED: Properly handle columnar format
            # ============================================================
            if isinstance(data, dict):
                # Check for columnar format (cols + data/rows)
                if 'cols' in data:
                    cols = data['cols']
                    
                    # Get the actual data rows
                    if 'data' in data:
                        rows = data['data']
                    elif 'rows' in data:
                        rows = data['rows']
                    else:
                        print(f"❌ 'cols' found but no 'data' or 'rows' key")
                        break
                    
                    if not rows:
                        print(f"ℹ️ No candidates found on page {current_page}, stopping")
                        break
                    
                    # Create DataFrame from columnar data
                    df_page = pd.DataFrame(rows, columns=cols)
                    
                    # Debug first page
                    if current_page == 1:
                        print(f"\n🔍 COLUMNAR FORMAT DETECTED:")
                        print(f"   Columns: {cols}")
                        print(f"   First row sample:")
                        if len(df_page) > 0:
                            first_row = df_page.iloc[0]
                            for col in cols[:8]:  # Show first 8 columns
                                print(f"      {col}: {first_row[col]}")
                    
                    all_candidates.append(df_page)
                    print(f"✅ Page {current_page}: {len(df_page)} candidates")
                    
                    if len(df_page) < page_size:
                        print(f"ℹ️ Got fewer candidates ({len(df_page)}) than page size ({page_size}), assuming last page")
                        break
                    
                # Check for object format (data/candidates/results as list of objects)
                elif 'data' in data:
                    candidates = data['data']
                    print(f"   Found 'data' key with {len(candidates)} items")
                elif 'candidates' in data:
                    candidates = data['candidates']
                    print(f"   Found 'candidates' key with {len(candidates)} items")
                elif 'results' in data:
                    candidates = data['results']
                    print(f"   Found 'results' key with {len(candidates)} items")
                else:
                    # Try to use entire dict as single candidate
                    candidates = [data]
                    print(f"⚠️ Using entire response as single candidate")
                
                # Process object format (only if we didn't already handle columnar format)
                if 'cols' not in data:
                    if not candidates:
                        print(f"ℹ️ No candidates found on page {current_page}, stopping")
                        break
                    
                    df_page = pd.DataFrame(candidates)
                    all_candidates.append(df_page)
                    
                    print(f"✅ Page {current_page}: Got {len(df_page)} candidates")
                    
                    if len(df_page) < page_size:
                        print(f"ℹ️ Got fewer candidates ({len(df_page)}) than page size ({page_size}), assuming last page")
                        break
                        
            elif isinstance(data, list):
                candidates = data
                df_page = pd.DataFrame(candidates)
                all_candidates.append(df_page)
                print(f"✅ Page {current_page}: Got {len(df_page)} candidates")
                
                if len(df_page) < page_size:
                    break
            else:
                print(f"❌ Unexpected API response format on page {current_page}")
                break
            
            current_page += 1
            
            import time
            time.sleep(0.1)
            
        except requests.exceptions.RequestException as e:
            print(f"❌ Error fetching page {current_page}: {str(e)}")
            break
        except json.JSONDecodeError as e:
            print(f"❌ JSON parsing error on page {current_page}: {str(e)}")
            break
        except Exception as e:
            print(f"❌ Unexpected error on page {current_page}: {str(e)}")
            import traceback
            traceback.print_exc()
            break
    
    if all_candidates:
        final_df = pd.concat(all_candidates, ignore_index=True)
        print(f"🎉 Fetching complete! Total candidates retrieved: {len(final_df)}")
        
        # Debug: Check skills column
        print(f"\n🔍 DATAFRAME ANALYSIS:")
        print(f"   Total rows: {len(final_df)}")
        print(f"   Columns: {final_df.columns.tolist()}")
        
        # Normalize the dataframe (this will map c_skills -> skills, etc.)
        normalized_df = normalize_candidate_dataframe(final_df)
        
        # Check if skills column exists after normalization
        if 'skills' in normalized_df.columns:
            print(f"\n✅ Skills column found after normalization!")
            print(f"   Non-null skills: {normalized_df['skills'].notna().sum()}")
            print(f"   Sample skills values:")
            for idx, skill_val in enumerate(normalized_df['skills'].head(5)):
                print(f"      [{idx}] {str(skill_val)[:200]}")
        else:
            print(f"\n❌ Skills column STILL not found after normalization!")
        
        return normalized_df
    else:
        print("⚠️ No candidates fetched from API")
        return pd.DataFrame()
    

def fetch_candidates_from_api_initial(api_url=None, timeout=30):
    '''Fetch initial candidate count without filters from Kunji API'''
    try:
        # Use base URL with minimal parameters for initial fetch
        if api_url is None:
            base_url = getattr(settings, 'CANDIDATES_API_BASE_URL', None)
            if not base_url:
                print("❌ Error: CANDIDATES_API_BASE_URL not configured in settings")
                return pd.DataFrame()
            
            # Simple query to get sample data
            api_url = f"{base_url}?page=1&limit=10"
        
        api_token = getattr(settings, 'CANDIDATES_API_TOKEN', None)
        
        headers = {'Content-Type': 'application/json'}
        if api_token:
            headers['Authorization'] = f'Bearer {api_token}'
        
        print(f"🔄 Fetching initial candidates from Kunji API: {api_url}")
        response = requests.get(api_url, headers=headers, timeout=timeout)
        response.raise_for_status()
        
        data = response.json()
        
        # Handle API response
        if isinstance(data, dict):
            if 'cols' in data and 'data' in data:
                df = pd.DataFrame(data['data'], columns=data['cols'])
            elif 'cols' in data and 'rows' in data:
                df = pd.DataFrame(data['rows'], columns=data['cols'])
            elif 'data' in data:
                df = pd.DataFrame(data['data'])
            else:
                df = pd.DataFrame([data])
        elif isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            print(f"❌ Unexpected API response format: {type(data)}")
            return pd.DataFrame()
        
        df = normalize_candidate_dataframe(df)
        print(f"✅ Successfully fetched {len(df)} initial candidates from Kunji API")
        return df
    
    except Exception as e:
        print(f"❌ Error fetching initial candidates: {e}")
        return pd.DataFrame()
    

def fetch_candidates_from_api(skills=None, experience=None, page=1, limit=100, timeout=30):
    '''
    Fetch candidate data from Kunji API with dynamic query parameters
    
    Args:
        skills: List of skills to filter by
        experience: Experience level to filter by
        page: Page number for pagination
        limit: Number of results per page
        timeout: Request timeout in seconds
    
    Returns:
        pandas DataFrame with candidate data
    '''
    try:
        # Build dynamic API URL
        api_url = build_api_url(skills=skills, experience=experience, page=page, limit=limit)
        api_token = getattr(settings, 'CANDIDATES_API_TOKEN', None)
        
        if not api_token:
            print("⚠️ Warning: CANDIDATES_API_TOKEN not configured in settings")
        
        headers = {'Content-Type': 'application/json'}
        if api_token:
            headers['Authorization'] = f'Bearer {api_token}'
        
        print(f"🔄 Fetching candidates from Kunji API: {api_url}")
        response = requests.get(api_url, headers=headers, timeout=timeout)
        response.raise_for_status()
        
        data = response.json()
        
        # Handle different API response structures
        if isinstance(data, dict):
            print(f"📦 API Response keys: {list(data.keys())}")
            
            if 'cols' in data and 'data' in data:
                print(f"✅ Using columnar format (cols + data)")
                df = pd.DataFrame(data['data'], columns=data['cols'])
                print(f"✅ Successfully fetched {len(df)} candidates from API")
                return normalize_candidate_dataframe(df)
            elif 'cols' in data and 'rows' in data:
                print(f"✅ Using columnar format (cols + rows)")
                df = pd.DataFrame(data['rows'], columns=data['cols'])
                print(f"✅ Successfully fetched {len(df)} candidates from API")
                return normalize_candidate_dataframe(df)
            elif 'data' in data:
                candidates = data['data']
                print(f"✅ Using 'data' key from response")
            elif 'candidates' in data:
                candidates = data['candidates']
                print(f"✅ Using 'candidates' key from response")
            elif 'results' in data:
                candidates = data['results']
                print(f"✅ Using 'results' key from response")
            else:
                # Try to use entire response as single candidate
                candidates = [data]
                print(f"⚠️ Using entire response as single candidate")
        elif isinstance(data, list):
            candidates = data
            print(f"✅ Response is a list with {len(data)} items")
        else:
            print(f"❌ Unexpected API response format: {type(data)}")
            return pd.DataFrame()
        
        if not candidates:
            print("⚠️ No candidates found in API response")
            return pd.DataFrame()
        
        print(f"📊 Creating DataFrame from {len(candidates)} candidates")
        df = pd.DataFrame(candidates)
        df = normalize_candidate_dataframe(df)
        
        print(f"✅ Successfully fetched {len(df)} candidates from Kunji API")
        return df
    
    except requests.exceptions.Timeout:
        print(f"❌ API request timed out after {timeout} seconds")
        return pd.DataFrame()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error fetching candidates from API: {e}")
        return pd.DataFrame()
    except json.JSONDecodeError as e:
        print(f"❌ Error parsing API response JSON: {e}")
        return pd.DataFrame()
    except Exception as e:
        print(f"❌ Unexpected error fetching candidates: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()


# ============================================================================
# REST OF THE CODE - MODIFIED FOR BATCH PROCESSING
# ============================================================================

def process_candidate_with_tokens(row, required_skills, priority_skills, nice_to_have_skills,
                                  jd_role_title, min_req_skills, min_quality_threshold, 
                                  min_match_percentage, location_preference, required_experience, 
                                  filtered_count, token_aggregator):
    '''Process a single candidate - OPTIMIZED with batch processing prep'''
    candidate_skills_str = str(row.get('skills', ''))
    
    if not candidate_skills_str or candidate_skills_str.lower() in ['nan', 'none', '']:
        return None
    
    # Parse and filter candidate skills
    candidate_skills = parse_candidate_skills(candidate_skills_str)
    if not candidate_skills:
        return None
    
    # Skill matching
    matched_details, total_weighted_score, total_weight, nice_to_have_matched = match_skills(
        candidate_skills, required_skills, priority_skills, nice_to_have_skills, jd_role_title
    )
    
    required_matched = sum(1 for s in required_skills if s in matched_details)
    
    if required_matched > 0:
        print(f"   Candidate: {row.get('name', 'Unknown')} - Matched {required_matched}/{len(required_skills)} skills")
    
    if required_matched < min_req_skills:
        filtered_count['min_skills'] += 1
        return None
    
    # NOTE: Skill relevance and designation similarity will be batch processed
    # For now, use cached values or defaults
    candidate_designation = str(row.get('designation', ''))
    matched_skills = list(matched_details.keys())
    
    skill_relevance_score = calculate_skill_relevance_score(
        matched_skills, jd_role_title, candidate_designation
    )
    
    # Calculate scores
    candidate_scores = calculate_candidate_scores(
        matched_details, required_skills, required_matched, 
        total_weighted_score, total_weight, skill_relevance_score,
        jd_role_title, candidate_designation, row, location_preference,
        required_experience, priority_skills, nice_to_have_matched, nice_to_have_skills
    )
    
    effective_quality_threshold = max(min_quality_threshold, 5)
    
    if candidate_scores['quality_score'] < effective_quality_threshold:
        filtered_count['quality'] += 1
        return None
    
    if candidate_scores['combined_score'] < min_match_percentage:
        filtered_count['threshold'] += 1
        return None
    
    # Build final candidate data
    return build_candidate_data(row, candidate_scores, matched_details, matched_skills, 
                               required_skills, priority_skills, nice_to_have_matched)


def match_candidates_with_jd(required_skills=['all_skills'], min_match_percentage=15, api_url=None, api_key=None,
                             priority_skills=None, nice_to_have_skills=None, use_fuzzy=True, 
                             location_preference=None, required_experience=None,
                             min_required_skills_match=None,
                             industry_preference=None, 
                             min_quality_threshold=5,
                             jd_role_title=None,
                             api_filter_skills=None,
                             max_candidates_from_api=500,
                             debug_mode=True):
    '''
    OPTIMIZED VERSION with better error handling for malformed data
    '''
    total_token_usage = {
        'prompt_tokens': 0,
        'completion_tokens': 0,
        'total_tokens': 0,
        'api_calls': 0
    }
    
    api_stats = {
        'total_fetched': 0,
        'pages_fetched': 0,
        'filter_skills': api_filter_skills or required_skills[:10],
        'filter_experience': required_experience,
        'candidates_with_skills': 0,
        'candidates_without_skills': 0
    }
    
    try:
        print(f"\n{'='*80}")
        print(f"🎯 FETCHING TOP {max_candidates_from_api} CANDIDATES FROM API")
        print(f"{'='*80}")
        
        filter_skills = api_filter_skills if api_filter_skills else required_skills[:10]
        
        print(f"📋 Filter Skills: {filter_skills}")
        print(f"💼 Filter Experience: {required_experience or 'Any'}")
        
        # Fetch candidates from API with filters
        df = fetch_all_candidates_from_api(
            skills=filter_skills,
            experience=required_experience,
            page_size=100,
            max_pages=5,
            timeout=30
        )
        
        api_stats['total_fetched'] = len(df)
        api_stats['pages_fetched'] = (len(df) // 100) + (1 if len(df) % 100 > 0 else 0)
        
        if df.empty:
            print("❌ No candidates found from API")
            return {
                'candidates': [],
                'token_usage': total_token_usage,
                'model': 'none',
                'api_stats': api_stats,
                'optimization_stats': {
                    'total_candidates': 0,
                    'matched_candidates': 0,
                    'api_calls': 0
                }
            }
        
        if 'skills' not in df.columns:
            print("❌ Error: 'skills' column not found")
            print(f"   Available columns: {df.columns.tolist()}")
            return {
                'candidates': [],
                'token_usage': total_token_usage,
                'model': 'none',
                'api_stats': api_stats,
                'optimization_stats': {
                    'total_candidates': len(df),
                    'matched_candidates': 0,
                    'api_calls': 0
                }
            }
        
        print(f"\n{'='*80}")
        print(f"📊 PROCESSING {len(df)} FILTERED CANDIDATES")
        print(f"{'='*80}")
        
        # Normalize inputs
        required_skills_lower = [normalize_skill(s) for s in required_skills if s.strip()]
        priority_skills_lower = [normalize_skill(s) for s in (priority_skills or [])]
        nice_to_have_lower = [normalize_skill(s) for s in (nice_to_have_skills or [])]
        
        if not required_skills_lower:
            print("❌ No valid required skills")
            return {
                'candidates': [],
                'token_usage': total_token_usage,
                'model': 'none',
                'api_stats': api_stats,
                'optimization_stats': {
                    'total_candidates': len(df),
                    'matched_candidates': 0,
                    'api_calls': 0
                }
            }
        
        # ADAPTIVE MINIMUM: Start with 2 skills if data quality is poor
        if min_required_skills_match is None:
            # Check data quality first
            sample_skills = df['skills'].head(20)
            valid_skills_count = sum(1 for s in sample_skills if parse_candidate_skills(str(s)))
            
            if valid_skills_count < 10:  # Less than 50% have valid skills
                print(f"⚠️ LOW DATA QUALITY DETECTED - Using minimum 2 skills instead of 4")
                min_req_skills = 2
            else:
                min_req_skills = 4
        else:
            min_req_skills = max(2, min_required_skills_match)
        
        print(f"🎯 Required skills in JD: {len(required_skills_lower)}")
        print(f"🎯 Minimum skills to match: {min_req_skills}")
        
        # Filter and score candidates
        print(f"\n🔍 Filtering and scoring candidates...")
        matched_candidates = []
        filtered_count = defaultdict(int)
        candidates_processed = 0
        candidates_with_valid_skills = 0
        
        for idx, row in df.iterrows():
            candidates_processed += 1
            
            candidate_skills_str = str(row.get('skills', ''))
            
            # Debug every 50th candidate
            if candidates_processed % 50 == 0:
                print(f"   Processed {candidates_processed}/{len(df)} candidates...")
            
            if not candidate_skills_str or candidate_skills_str.lower().strip() in ['nan', 'none', 'n/a', 'n a', '']:
                filtered_count['no_skills_data'] += 1
                continue
            
            candidate_skills = parse_candidate_skills(candidate_skills_str)
            
            if not candidate_skills:
                filtered_count['invalid_skills'] += 1
                continue
            
            candidates_with_valid_skills += 1
            
            # Show first few candidates with valid skills
            if candidates_with_valid_skills <= 3:
                print(f"\n   ✓ Candidate #{candidates_with_valid_skills}: {row.get('name', 'Unknown')}")
                print(f"      Skills: {candidate_skills[:5]}...")
            
            # Skill matching
            matched_details, total_weighted_score, total_weight, nice_to_have_matched = match_skills(
                candidate_skills, required_skills_lower, priority_skills_lower, nice_to_have_lower, jd_role_title
            )
            
            required_matched = sum(1 for s in required_skills_lower if s in matched_details)
            
            if required_matched < min_req_skills:
                filtered_count['min_skills'] += 1
                continue
            
            # Show matches
            if len(matched_candidates) < 5:
                print(f"\n   🎯 MATCH FOUND: {row.get('name', 'Unknown')}")
                print(f"      Matched {required_matched}/{len(required_skills_lower)} skills")
                print(f"      Skills: {list(matched_details.keys())[:5]}")
            
            candidate_designation = str(row.get('designation', ''))
            matched_skills = list(matched_details.keys())
            
            skill_relevance_score = min(100, (required_matched / len(required_skills_lower)) * 120)
            
            candidate_scores = calculate_candidate_scores_fast(
                matched_details, required_skills_lower, required_matched, 
                total_weighted_score, total_weight, skill_relevance_score,
                jd_role_title, candidate_designation, row, location_preference,
                required_experience, priority_skills_lower, nice_to_have_matched, nice_to_have_lower
            )
            
            effective_quality_threshold = max(min_quality_threshold, 5)
            
            if candidate_scores['quality_score'] < effective_quality_threshold:
                filtered_count['quality'] += 1
                continue
            
            if candidate_scores['combined_score'] < min_match_percentage:
                filtered_count['threshold'] += 1
                continue
            
            final_candidate = build_candidate_data(
                row, candidate_scores, matched_details, matched_skills, 
                required_skills_lower, priority_skills_lower, nice_to_have_matched
            )
            matched_candidates.append(final_candidate)
        
        api_stats['candidates_with_skills'] = candidates_with_valid_skills
        api_stats['candidates_without_skills'] = len(df) - candidates_with_valid_skills
        
        # Sort results
        result = finalize_results(matched_candidates, filtered_count, len(df), min_match_percentage, min_req_skills)
        
        print(f"\n{'='*80}")
        print(f"📊 MATCHING SUMMARY")
        print(f"{'='*80}")
        print(f"   API candidates fetched: {api_stats['total_fetched']}")
        print(f"   Candidates with valid skills: {candidates_with_valid_skills}")
        print(f"   Candidates without skills: {api_stats['candidates_without_skills']}")
        print(f"   Pages fetched: {api_stats['pages_fetched']}")
        print(f"   Candidates matched: {len(matched_candidates)}")
        print(f"   Minimum skills required: {min_req_skills}")
        print(f"{'='*80}\n")
        
        return {
            'candidates': result,
            'token_usage': total_token_usage,
            'model': 'none',
            'api_stats': api_stats,
            'optimization_stats': {
                'total_candidates': len(df),
                'candidates_with_valid_skills': candidates_with_valid_skills,
                'matched_candidates': len(matched_candidates),
                'api_calls': 0
            }
        }
    
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return {
            'candidates': [],
            'token_usage': total_token_usage,
            'model': 'none',
            'api_stats': api_stats,
            'optimization_stats': {
                'total_candidates': 0,
                'matched_candidates': 0,
                'api_calls': 0
            }
        }

# ============================================================================
# ALL OTHER FUNCTIONS REMAIN UNCHANGED
# ============================================================================

def calculate_min_required_skills(jd_role_title, required_skills, min_override):
    '''Calculate minimum required skills - DEFAULT: 4 skills minimum'''
    if min_override is not None:
        return max(4, min_override)
    return 4

def process_candidate(row, required_skills, priority_skills, nice_to_have_skills,
                     jd_role_title, min_req_skills, min_quality_threshold, 
                     min_match_percentage, location_preference, required_experience, filtered_count):
    '''Process a single candidate - ULTRA LENIENT VERSION'''
    candidate_skills_str = str(row.get('skills', ''))
    
    if not candidate_skills_str or candidate_skills_str.lower() in ['nan', 'none', '']:
        return None
    
    # REMOVED: Designation filter - accept all designations
    
    # Parse and filter candidate skills
    candidate_skills = parse_candidate_skills(candidate_skills_str)
    if not candidate_skills:
        return None
    
    # Skill matching with ULTRA LOW thresholds
    matched_details, total_weighted_score, total_weight, nice_to_have_matched = match_skills(
        candidate_skills, required_skills, priority_skills, nice_to_have_skills, jd_role_title
    )
    
    required_matched = sum(1 for s in required_skills if s in matched_details)
    
    # Debug print
    if required_matched > 0:
        print(f"   Candidate: {row.get('name', 'Unknown')} - Matched {required_matched}/{len(required_skills)} skills")
    
    # CHANGED: Accept if ANY skill matches
    if required_matched < min_req_skills:
        filtered_count['min_skills'] += 1
        return None
    
    # REMOVED: Skill relevance threshold - accept all candidates with any match
    candidate_designation = str(row.get('designation', ''))
    matched_skills = list(matched_details.keys())
    skill_relevance_score = calculate_skill_relevance_score(
        matched_skills, jd_role_title, candidate_designation
    )
    
    # Calculate scores
    candidate_scores = calculate_candidate_scores(
        matched_details, required_skills, required_matched, 
        total_weighted_score, total_weight, skill_relevance_score,
        jd_role_title, candidate_designation, row, location_preference,
        required_experience, priority_skills, nice_to_have_matched, nice_to_have_skills
    )
    
    # CHANGED: Ultra low quality threshold of 5
    effective_quality_threshold = max(min_quality_threshold, 5)
    
    if candidate_scores['quality_score'] < effective_quality_threshold:
        filtered_count['quality'] += 1
        return None
    
    # CHANGED: Ultra low match percentage threshold
    if candidate_scores['combined_score'] < min_match_percentage:
        filtered_count['threshold'] += 1
        return None
    
    # Build final candidate data
    return build_candidate_data(row, candidate_scores, matched_details, matched_skills, 
                               required_skills, priority_skills, nice_to_have_matched)

def is_candidate_designation_relevant(jd_role_title, candidate_designation):
    '''ULTRA LENIENT: Always return True - accept all designations'''
    return True

def parse_candidate_skills(skills_str):
    '''Parse and filter candidate skills - ROBUST VERSION with better error handling'''
    if not skills_str or str(skills_str).lower().strip() in ['nan', 'none', 'n/a', 'n a', '']:
        return []
    
    # Convert to string and clean
    skills_str = str(skills_str).strip()
    
    # Check if it's a JSON array string
    if skills_str.startswith('[') and skills_str.endswith(']'):
        try:
            import json
            skills_list = json.loads(skills_str)
            if isinstance(skills_list, list):
                return [normalize_skill(s) for s in skills_list if s and len(str(s).strip()) > 2]
        except:
            pass
    
    # Split by common delimiters
    candidate_skills_raw = re.split(r'[,;|\n]+', skills_str)
    candidate_skills = []
    
    for skill in candidate_skills_raw:
        skill = skill.strip()
        
        # Skip invalid skills
        if not skill or len(skill) <= 2:
            continue
        
        # Skip common invalid values
        if skill.lower() in ['nan', 'none', 'n/a', 'n a', 'na', 'null', '']:
            continue
        
        # Normalize and add
        normalized = normalize_skill(skill)
        if normalized and len(normalized) > 2:
            candidate_skills.append(normalized)
    
    return candidate_skills

def match_skills(candidate_skills, required_skills, priority_skills, nice_to_have_skills, jd_role_title):
    '''
    Match candidate skills with SMART MATCHING
    - Accepts scores >= 55 (partial matches)
    - Uses token-based matching for multi-word skills
    '''
    matched_details = {}
    total_weighted_score = 0
    total_weight = 0
    
    # Normalize candidate skills once
    candidate_skills_normalized = [normalize_skill(s) for s in candidate_skills]
    
    print(f"🔍 Matching {len(required_skills)} JD skills against {len(candidate_skills_normalized)} candidate skills")
    
    # LOWERED THRESHOLD: Accept scores >= 55 (was 100 in exact matching)
    MATCH_THRESHOLD = 55
    
    # ========================================================================
    # 1. MATCH REQUIRED SKILLS (Weight: 1.0)
    # ========================================================================
    print(f"\n📋 Checking {len(required_skills)} required skills...")
    
    for req_skill in required_skills:
        best_score = 0
        best_match = None
        best_cand_skill = None
        
        # Find best match among all candidate skills
        for i, cand_skill in enumerate(candidate_skills):
            score, match_type, details = calculate_skill_similarity(req_skill, cand_skill, True)
            
            if score > best_score:
                best_score = score
                best_match = {'type': match_type, 'details': details}
                best_cand_skill = cand_skill
        
        if best_score >= MATCH_THRESHOLD:
            matched_details[req_skill] = {
                'score': best_score,
                'weight': 1.0,
                'category': 'required',
                'cand_skill': best_cand_skill,
                **best_match
            }
            total_weighted_score += best_score * 1.0
            
            # Show match type
            if best_score == 100:
                print(f"   ✅ {req_skill} → {best_cand_skill}")
            elif best_score >= 80:
                print(f"   ⚡ {req_skill} ≈ {best_cand_skill} ({best_score})")
            else:
                print(f"   💡 {req_skill} ~ {best_cand_skill} ({best_score})")
        else:
            print(f"   ❌ {req_skill}")
        
        total_weight += 1.0
    
    # ========================================================================
    # 2. MATCH PRIORITY SKILLS (Weight: 2.0)
    # ========================================================================
    if priority_skills:
        print(f"\n⭐ Checking {len(priority_skills)} priority skills...")
        
        for pri_skill in priority_skills:
            # Upgrade if already matched
            if pri_skill in matched_details:
                matched_details[pri_skill]['weight'] = 2.0
                matched_details[pri_skill]['category'] = 'priority'
                total_weighted_score += matched_details[pri_skill]['score'] * 1.0
                total_weight += 1.0
                print(f"   ⭐ Upgraded: {pri_skill}")
                continue
            
            # Find best match
            best_score = 0
            best_match = None
            best_cand_skill = None
            
            for cand_skill in candidate_skills:
                score, match_type, details = calculate_skill_similarity(pri_skill, cand_skill, True)
                if score > best_score:
                    best_score = score
                    best_match = {'type': match_type, 'details': details}
                    best_cand_skill = cand_skill
            
            if best_score >= MATCH_THRESHOLD:
                matched_details[pri_skill] = {
                    'score': best_score,
                    'weight': 2.0,
                    'category': 'priority',
                    'cand_skill': best_cand_skill,
                    **best_match
                }
                total_weighted_score += best_score * 2.0
                total_weight += 2.0
                print(f"   ✅ Priority: {pri_skill} → {best_cand_skill} ({best_score})")
            else:
                total_weight += 2.0
                print(f"   ❌ Priority: {pri_skill}")
    
    # ========================================================================
    # 3. MATCH NICE-TO-HAVE SKILLS (Weight: 0.5)
    # ========================================================================
    nice_to_have_matched = 0
    
    if nice_to_have_skills:
        print(f"\n💡 Checking {len(nice_to_have_skills)} nice-to-have skills...")
        
        for nice_skill in nice_to_have_skills:
            if nice_skill in matched_details:
                continue
            
            best_score = 0
            best_match = None
            best_cand_skill = None
            
            for cand_skill in candidate_skills:
                score, match_type, details = calculate_skill_similarity(nice_skill, cand_skill, True)
                if score > best_score:
                    best_score = score
                    best_match = {'type': match_type, 'details': details}
                    best_cand_skill = cand_skill
            
            if best_score >= MATCH_THRESHOLD:
                matched_details[nice_skill] = {
                    'score': best_score,
                    'weight': 0.5,
                    'category': 'nice_to_have',
                    'cand_skill': best_cand_skill,
                    **best_match
                }
                nice_to_have_matched += 1
                total_weighted_score += best_score * 0.5
                print(f"   ✅ Bonus: {nice_skill} → {best_cand_skill}")
    
    # ========================================================================
    # SUMMARY
    # ========================================================================
    required_matched = sum(1 for d in matched_details.values() if d['category'] in ['required', 'priority'])
    
    print(f"\n📊 Match Summary:")
    print(f"   Required/Priority: {required_matched}/{len(required_skills) + len(priority_skills or [])}")
    print(f"   Nice-to-have: {nice_to_have_matched}/{len(nice_to_have_skills or [])}")
    print(f"   Total matched: {len(matched_details)}")
    print(f"   Weighted score: {total_weighted_score:.1f}/{total_weight:.1f}")
    
    return matched_details, total_weighted_score, total_weight, nice_to_have_matched

def find_best_skill_match(target_skill, candidate_skills, jd_role_title):
    '''Find exact match only'''
    target_norm = normalize_skill(target_skill)
    
    for cand_skill in candidate_skills:
        cand_norm = normalize_skill(cand_skill)
        
        if target_norm == cand_norm:
            return 100, {'type': 'exact', 'cand_skill': cand_skill, 'details': {}}
        
        # Optional: substring match
        if target_norm in cand_norm or cand_norm in target_norm:
            return 85, {'type': 'substring', 'cand_skill': cand_skill, 'details': {}}
    
    return 0, None

def calculate_candidate_scores_fast(matched_details, required_skills, required_matched, 
                                    total_weighted_score, total_weight, skill_relevance_score,
                                    jd_role_title, candidate_designation, row, location_preference,
                                    required_experience, priority_skills, nice_to_have_matched, nice_to_have_skills):
    '''Calculate all scores for a candidate - FAST VERSION (No OpenAI calls)'''
    base_match_pct = (required_matched / len(required_skills)) * 100 if required_skills else 0
    quality_score = (total_weighted_score / total_weight) if total_weight > 0 else 0
    
    # Use skill relevance score directly (already calculated based on match percentage)
    relevance_multiplier = max(0.85, skill_relevance_score / 100)
    quality_score = quality_score * relevance_multiplier
    
    # Combined score
    combined_score = (base_match_pct * 0.6) + (quality_score * 0.4)
    
    # Apply bonuses (without OpenAI calls)
    bonuses = calculate_bonuses_fast(
        jd_role_title, candidate_designation, skill_relevance_score,
        row, location_preference, required_experience, priority_skills,
        nice_to_have_matched, nice_to_have_skills
    )
    
    combined_score += sum(bonuses.values())
    combined_score = min(combined_score, 100)
    
    return {
        'base_match_pct': base_match_pct,
        'quality_score': quality_score,
        'combined_score': combined_score,
        'skill_relevance_score': skill_relevance_score,
        'bonuses': bonuses
    }

def calculate_bonuses_fast(jd_role_title, candidate_designation, skill_relevance_score,
                           row, location_preference, required_experience, priority_skills,
                           nice_to_have_matched, nice_to_have_skills):
    '''Calculate all bonus points for a candidate - FAST VERSION (No OpenAI calls)'''
    bonuses = {}
    
    # Designation bonus using fuzzy matching only (no OpenAI)
    if jd_role_title and candidate_designation:
        jd_clean = str(jd_role_title).lower().strip()
        cand_clean = str(candidate_designation).lower().strip()
        
        if jd_clean and cand_clean and cand_clean not in ['nan', 'none', 'n/a', '']:
            # Use fuzzy matching
            token_sort = fuzz.token_sort_ratio(jd_clean, cand_clean)
            if token_sort >= 60:
                desg_bonus = (token_sort / 100) * 15
                bonuses['designation'] = round(desg_bonus, 1)
    
    if skill_relevance_score >= 50:
        relevance_bonus = ((skill_relevance_score - 50) / 50) * 10
        bonuses['skill_relevance'] = round(relevance_bonus, 1)
    
    if priority_skills:
        priority_matched = sum(1 for s in priority_skills if s in [k for k in row.keys()])
        priority_pct = (priority_matched / len(priority_skills))
        priority_bonus = priority_pct * 15
        bonuses['priority'] = round(priority_bonus, 1)
    
    if nice_to_have_skills:
        nice_bonus = (nice_to_have_matched / len(nice_to_have_skills)) * 5
        bonuses['nice_to_have'] = round(nice_bonus, 1)
    
    if required_experience:
        exp_bonus = calculate_experience_score(row.get('experience'), required_experience) * (10/15)
        if exp_bonus > 0:
            bonuses['experience'] = round(exp_bonus, 1)
    
    if location_preference:
        locations = location_preference if isinstance(location_preference, list) else [location_preference]
        cand_location = str(row.get('location', '')).lower()
        for loc in locations:
            if loc.lower() in cand_location:
                bonuses['location'] = 5
                break
    
    return bonuses

def calculate_candidate_scores(matched_details, required_skills, required_matched, 
                              total_weighted_score, total_weight, skill_relevance_score,
                              jd_role_title, candidate_designation, row, location_preference,
                              required_experience, priority_skills, nice_to_have_matched, nice_to_have_skills):
    '''Calculate all scores for a candidate'''
    base_match_pct = (required_matched / len(required_skills)) * 100 if required_skills else 0
    quality_score = (total_weighted_score / total_weight) if total_weight > 0 else 0
    
    # CHANGED: More lenient relevance multiplier (minimum 0.85 instead of 0.7)
    relevance_multiplier = max(0.85, skill_relevance_score / 100)
    quality_score = quality_score * relevance_multiplier
    
    # Combined score
    combined_score = (base_match_pct * 0.6) + (quality_score * 0.4)
    
    # Apply bonuses
    bonuses = calculate_bonuses(
        jd_role_title, candidate_designation, skill_relevance_score,
        row, location_preference, required_experience, priority_skills,
        nice_to_have_matched, nice_to_have_skills
    )
    
    combined_score += sum(bonuses.values())
    combined_score = min(combined_score, 100)
    
    return {
        'base_match_pct': base_match_pct,
        'quality_score': quality_score,
        'combined_score': combined_score,
        'skill_relevance_score': skill_relevance_score,
        'bonuses': bonuses
    }

def calculate_bonuses(jd_role_title, candidate_designation, skill_relevance_score,
                     row, location_preference, required_experience, priority_skills,
                     nice_to_have_matched, nice_to_have_skills):
    '''Calculate all bonus points for a candidate'''
    bonuses = {}
    
    if jd_role_title:
        designation_score, _, designation_details = calculate_designation_similarity(
            jd_role_title, candidate_designation, True
        )
        if designation_score > 0:
            desg_bonus = (designation_score / 100) * 15
            bonuses['designation'] = round(desg_bonus, 1)
    
    if skill_relevance_score >= 50:
        relevance_bonus = ((skill_relevance_score - 50) / 50) * 10
        bonuses['skill_relevance'] = round(relevance_bonus, 1)
    
    if priority_skills:
        priority_matched = sum(1 for s in priority_skills if s in [k for k in row.keys()])
        priority_pct = (priority_matched / len(priority_skills))
        priority_bonus = priority_pct * 15
        bonuses['priority'] = round(priority_bonus, 1)
    
    if nice_to_have_skills:
        nice_bonus = (nice_to_have_matched / len(nice_to_have_skills)) * 5
        bonuses['nice_to_have'] = round(nice_bonus, 1)
    
    if required_experience:
        exp_bonus = calculate_experience_score(row.get('experience'), required_experience) * (10/15)
        if exp_bonus > 0:
            bonuses['experience'] = round(exp_bonus, 1)
    
    if location_preference:
        locations = location_preference if isinstance(location_preference, list) else [location_preference]
        cand_location = str(row.get('location', '')).lower()
        for loc in locations:
            if loc.lower() in cand_location:
                bonuses['location'] = 5
                break
    
    return bonuses

def build_candidate_data(row, scores, matched_details, matched_skills, 
                        required_skills, priority_skills, nice_to_have_matched):
    '''Build the final candidate data dictionary'''
    exact_matches = sum(1 for d in matched_details.values() if d['type'] == 'exact')
    priority_matches = sum(1 for s in priority_skills if s in matched_details)
    skill_scores = [d['score'] for d in matched_details.values()]
    avg_skill_strength = np.mean(skill_scores) if skill_scores else 0
    
    # Use fuzzy matching for designation (no OpenAI)
    jd_role = str(scores.get('jd_role_title', '')).lower().strip()
    cand_desg = str(row.get('designation', '')).lower().strip()
    
    if jd_role and cand_desg and cand_desg not in ['nan', 'none', 'n/a', '']:
        designation_score = fuzz.token_sort_ratio(jd_role, cand_desg)
        if designation_score >= 80:
            designation_match_type = "high"
        elif designation_score >= 60:
            designation_match_type = "moderate"
        else:
            designation_match_type = "low"
        designation_reasoning = f"Fuzzy match: {designation_score}%"
    else:
        designation_score = 0
        designation_match_type = "no_data"
        designation_reasoning = "N/A"
    
    return {
        'id': row.get('id', 'N/A'),
        'name': row.get('name', 'N/A'),
        'email': row.get('email', 'N/A'),
        'contact': row.get('contact', 'N/A'),
        'location': row.get('location', 'N/A'),
        'current_company': row.get('current_company', 'N/A'),
        'designation': row.get('designation', 'N/A'),
        'experience': row.get('experience', 'N/A'),
        'linkedin': row.get('linkedin', 'N/A'),
        'qualification': row.get('qualification', 'N/A'),
        'skills': row.get('skills', 'N/A'),
        'cv_link': row.get('cv_link', 'N/A'),
        'status': row.get('status', 'Active'),
        'match_percentage': round(scores['combined_score'], 1),
        'quality_score': round(scores['quality_score'], 1),
        'base_match_percentage': round(scores['base_match_pct'], 1),
        'avg_skill_strength': round(avg_skill_strength, 1),
        'skill_relevance_score': round(scores['skill_relevance_score'], 1),
        'designation_match_score': round(designation_score, 1),
        'designation_match_type': designation_match_type,
        'designation_reasoning': designation_reasoning,
        'designation_confidence': 'fuzzy_match',
        'seniority_match': False,
        'function_match': False,
        'role_equivalent': False,
        'matched_skills': matched_skills,
        'matched_skills_count': len(matched_skills),
        'total_required_skills': len(required_skills),
        'exact_matches': exact_matches,
        'priority_matches': priority_matches,
        'nice_to_have_matches': nice_to_have_matched,
        'skill_match_details': matched_details,
        'bonuses': scores['bonuses'],
        'total_bonus': round(sum(scores['bonuses'].values()), 1)
    }

def finalize_results(matched_candidates, filtered_count, total_candidates, min_match_percentage, min_req_skills):
    '''Sort results and display summary'''
    # Sort candidates
    matched_candidates.sort(
        key=lambda x: (
            x['match_percentage'],
            x['matched_skills_count'],
            x['skill_relevance_score'],
            x['designation_match_score'],
            x['quality_score']
        ),
        reverse=True
    )
    
    # Print summary
    print(f"\n📊 Filtering Summary:")
    print(f"   Total candidates: {total_candidates}")
    for filter_type, count in filtered_count.items():
        print(f"   Filtered by {filter_type}: {count}")
    print(f"\n✅ Found {len(matched_candidates)} matching candidates (Minimum {min_req_skills} skills required)")
    
    if matched_candidates:
        print(f"\n🏆 Top 10 Candidates:")
        for i, c in enumerate(matched_candidates[:10], 1):
            print(f"\n{i}. {c['name']} - {c['designation']}")
            print(f"   📧 {c['email']}")
            print(f"   📍 {c['location']} | 💼 {c['experience']}")
            print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            print(f"   Overall Match: {c['match_percentage']:.1f}%")
            print(f"   Matched Skills: {c['matched_skills_count']}/{c['total_required_skills']}")
            print(f"   Skills: {', '.join(c['matched_skills'][:8])}")
            if len(c['matched_skills']) > 8:
                print(f"          ... and {len(c['matched_skills']) - 8} more")
    else:
        print(f"\n⚠️ No candidates found with {min_req_skills}+ matching skills.")
    
    return matched_candidates

def debug_skill_matching(required_skills, candidate_skills_sample):
    '''Debug helper to see how skills are being normalized and matched'''
    print("\n🔍 SKILL MATCHING DEBUG")
    print("=" * 60)
    
    print("\n📋 Required Skills (after normalization):")
    for skill in required_skills[:10]:
        normalized = normalize_skill(skill)
        print(f"   '{skill}' → '{normalized}'")
    
    print("\n📋 Sample Candidate Skills (after normalization):")
    for skill in candidate_skills_sample[:10]:
        normalized = normalize_skill(skill)
        print(f"   '{skill}' → '{normalized}'")
    
    print("\n🎯 Testing Matches:")
    for req in required_skills[:5]:
        print(f"\n   Looking for: '{req}'")
        for cand in candidate_skills_sample[:10]:
            score, match_type, details = calculate_skill_similarity(req, cand)
            if score > 0:
                print(f"      ✓ Matched '{cand}' (score: {score}, type: {match_type})")
    
    print("\n" + "=" * 60)


def normalize_candidate_dataframe(df):
    '''Normalize candidate DataFrame columns to standard format - ENHANCED'''
    print(f"📋 Original columns: {df.columns.tolist()}")
    
    # Column mapping from API format to standard format
    column_mapping = {
        'c_id': 'id',
        'c_name': 'name',
        'c_email': 'email',
        'c_phone': 'contact',
        'c_loc': 'location',
        'c_l_url': 'linkedin',
        'c_exp': 'experience',
        'c_skills': 'skills',
        'c_qualifications': 'qualification',
        'c_designation': 'designation',
        'c_cv_url': 'cv_link',
        'c_company': 'current_company'
    }
    
    # Apply column renaming
    rename_dict = {old: new for old, new in column_mapping.items() if old in df.columns}
    if rename_dict:
        df = df.rename(columns=rename_dict)
        print(f"✅ Renamed columns: {rename_dict}")
    
    # Define standard columns we need
    standard_columns = [
        'id', 'name', 'email', 'contact', 'location', 
        'linkedin', 'experience', 'skills', 'qualification', 
        'designation', 'cv_link'
    ]
    
    # Add missing standard columns with 'N/A' default
    for col in standard_columns:
        if col not in df.columns:
            df[col] = 'N/A'
            print(f"⚠️ Added missing column: {col}")
    
    # Add additional columns if not present
    if 'current_company' not in df.columns:
        df['current_company'] = 'N/A'
    if 'status' not in df.columns:
        df['status'] = 'Active'
    
    print(f"✅ Final columns: {df.columns.tolist()}")
    
    # Debug: Show sample data
    if not df.empty:
        print(f"📊 Sample data (first row):")
        sample = df.iloc[0]
        print(f"   - Name: {sample.get('name', 'N/A')}")
        print(f"   - Email: {sample.get('email', 'N/A')}")
        print(f"   - Skills: {str(sample.get('skills', 'N/A'))[:100]}...")
        print(f"   - Designation: {sample.get('designation', 'N/A')}")
    
    return df

def save_jd_to_excel(jd_data):
    '''Save job description data to Excel database'''
    try:
        excel_path = Path(settings.EXCEL_DATABASE_PATH)
        data_dir = excel_path.parent
        
        df_new = pd.DataFrame([jd_data])
        
        if excel_path.exists():
            try:
                df_existing = pd.read_excel(excel_path)
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            except Exception as e:
                print(f"⚠️ Error reading existing Excel file: {e}, creating new one")
                df_combined = df_new
        else:
            df_combined = df_new
        
        data_dir.mkdir(parents=True, exist_ok=True)
        df_combined.to_excel(excel_path, index=False, engine='openpyxl')
        print(f"✅ Successfully saved JD data to Excel: {excel_path}")
        
    except Exception as e:
        print(f"❌ Error saving JD data to Excel: {e}")

def export_matched_candidates(request, matched_candidates, jd_pk=None):
    """Export matched candidates to Excel and store in Django session (Vercel compatible)"""
    try:
        if not matched_candidates:
            return False, "No candidates available to export."

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matched Candidates"

        priority_headers = [
            'name', 'email', 'contact', 'designation', 'current_company',
            'experience', 'location', 'match_percentage', 'matched_skills_count',
            'total_required_skills', 'matched_skills', 'quality_score',
            'skill_relevance_score', 'linkedin', 'cv_link'
        ]
        
        all_keys = set()
        for candidate in matched_candidates:
            all_keys.update(candidate.keys())
        
        headers = []
        for h in priority_headers:
            if h in all_keys:
                headers.append(h)
        
        for key in sorted(all_keys):
            if key not in headers:
                headers.append(key)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, candidate in enumerate(matched_candidates, start=2):
            for col_index, key in enumerate(headers, start=1):
                value = candidate.get(key, "N/A")
                
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif isinstance(value, dict):
                    value = str(value)
                
                ws.cell(row=row_index, column=col_index, value=value)

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file_bytes = buffer.read()
        file_b64 = base64.b64encode(file_bytes).decode("utf-8")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"matched_candidates_{jd_pk or 'export'}_{timestamp}.xlsx"

        request.session["excel_file_data"] = file_b64
        request.session["excel_filename"] = filename
        
        file_size_kb = len(file_bytes) / 1024
        
        logger.info(f"✅ Excel export stored in session: {filename} ({file_size_kb:.2f} KB)")
        logger.info(f"   - Candidates: {len(matched_candidates)}")
        logger.info(f"   - Columns: {len(headers)}")
        logger.info(f"   - Base64 length: {len(file_b64)}")

        return True, f"Excel file generated successfully! ({len(matched_candidates)} candidates)"

    except Exception as e:
        logger.error(f"❌ Error exporting candidates: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, f"Error exporting candidates: {str(e)}"
    
def cleanup_old_matched_files(days=1):
    '''Delete matched candidate files older than specified days'''
    from datetime import datetime, timedelta
    
    matched_dir = Path(settings.MEDIA_ROOT) / "matched_candidates"
    if not matched_dir.exists():
        return
    
    cutoff_time = datetime.now() - timedelta(days=days)
    deleted_count = 0
    
    for file_path in matched_dir.glob('*.xlsx'):
        if datetime.fromtimestamp(file_path.stat().st_mtime) < cutoff_time:
            try:
                file_path.unlink()
                deleted_count += 1
                print(f"✅ Deleted old file: {file_path.name}")
            except Exception as e:
                print(f"⚠️ Could not delete {file_path.name}: {e}")
    
    if deleted_count > 0:
        print(f"✅ Cleanup complete: {deleted_count} old files deleted")



def delete_file_after_delay(file_path, delay_seconds=5):
    '''Delete a file after delay in background thread'''
    import time
    import threading
    
    def delete_file():
        if delay_seconds > 0:
            time.sleep(delay_seconds)
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"✅ Deleted file: {file_path}")
        except Exception as e:
            print(f"⚠️ Could not delete file {file_path}: {e}")
    
    if delay_seconds > 0:
        thread = threading.Thread(target=delete_file)
        thread.daemon = True
        thread.start()
    else:
        delete_file()


def extract_text_from_file(file_path):
    '''Extract text from TXT, PDF, or DOCX files'''
    ext = Path(file_path).suffix.lower()
    
    try:
        if ext == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        
        elif ext == '.pdf':
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            return text
        
        elif ext == '.docx':
            doc = Document(file_path)
            return '\n'.join([para.text for para in doc.paragraphs])
    
    except Exception as e:
        print(f"Error extracting text from {file_path}: {e}")
        return ""
    
    return ""

def generate_linkedin_search_strings(skills, role_title, experience_level):
    '''Generate optimized LinkedIn Recruiter boolean search strings'''
    
    top_skills = skills[:15] if len(skills) > 15 else skills
    
    searches = {}
    
    basic_and = " AND ".join([f'"{skill}"' for skill in top_skills[:8]])
    searches['basic_and'] = basic_and
    
    if len(top_skills) >= 3:
        part1 = " OR ".join([f'"{skill}"' for skill in top_skills[:3]])
        part2 = " OR ".join([f'"{skill}"' for skill in top_skills[3:6]])
        flexible = f'({part1}) AND ({part2})' if part2 else f'({part1})'
        searches['flexible'] = flexible
    
    skills_part = " AND ".join([f'"{skill}"' for skill in top_skills[:5]])
    title_search = f'(title:"{role_title}") AND ({skills_part})'
    searches['with_title'] = title_search
    
    skills_filter = ", ".join(top_skills[:10])
    searches['skills_filter'] = skills_filter
    
    xray_skills = " ".join([f'"{skill}"' for skill in top_skills[:6]])
    xray_search = f'site:linkedin.com/in/ "{role_title}" {xray_skills}'
    searches['xray'] = xray_search
    
    return searches

def extract_skills_from_jd(jd_text, domain_hint=""):
    '''Extract ALL skills comprehensively from job description using OpenAI API'''
    try:
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
    except Exception as e:
        print(f"Error initializing OpenAI client: {e}")
        return get_default_error_response()
    
    domain_context = f"The job is in the {domain_hint} domain." if domain_hint else ""
    
    prompt = f'''
You are an expert HR recruitment assistant AI. Carefully analyze the following Job Description and extract EVERY skill, technology, tool, qualification, and competency mentioned.

Return a structured JSON with:

1. "all_skills": A comprehensive list of ALL skills mentioned in the JD including:
   - Technical skills (programming languages, tools, frameworks, technologies)
   - Functional skills (domain-specific abilities)
   - Software/Tools (any applications, platforms, or systems)
   - Methodologies (Agile, Scrum, Six Sigma, etc.)
   - Certifications or qualifications mentioned
   - Domain knowledge areas
   - Soft skills (communication, leadership, teamwork, etc.)
   
   Extract 15-30 skills. Be thorough and don't miss anything mentioned in the JD.

2. "priority_skills": List 3-8 CRITICAL must-have skills that are absolutely essential for this role. These should be the skills mentioned multiple times or emphasized as requirements.

3. "nice_to_have_skills": List 5-10 bonus skills that would be beneficial but not mandatory.

4. "skill_categories": Organize the skills into categories like:
   {{"Technical": [...], "Tools": [...], "Soft Skills": [...], "Domain Knowledge": [...], "Certifications": [...]}}
   
5. "linkedin_optimized_skills": A list of 8-15 MOST IMPORTANT skills optimized for LinkedIn Recruiter search. 
   - Focus on searchable, industry-standard terms
   - Remove generic terms like "communication" or "teamwork"
   - Prioritize: specific technologies, tools, certifications, frameworks
   - Use exact names as they appear on LinkedIn (e.g., "JavaScript" not "JS", "Amazon Web Services (AWS)" not just "AWS")

6. "role_category": The most suitable role category (e.g., HR, Marketing, IT, Finance, Sales, Operations, etc.)

7. "experience_level": one of ["Entry Level", "Mid Level", "Senior Level", "Executive Level"]

8. "required_experience_years": Extract the experience requirement as a JSON object like {{"min": 3, "max": 5}} or {{"min": 5}} if only minimum is mentioned.

9. "key_responsibilities": List 5-7 main responsibilities mentioned in the JD

10. "qualifications": Educational requirements and certifications

11. "preferred_location": Extract any location preferences mentioned

{domain_context}

Be extremely thorough. If someone reads only your extracted skills, they should fully understand what this job requires.

Return ONLY valid JSON, no code block, no markdown, no explanation.

JD:
{jd_text[:4000]}
'''

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are an AI expert at extracting comprehensive skill requirements from job descriptions. Extract EVERY skill mentioned. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=2000
        )

        content = response.choices[0].message.content.strip()
        cleaned = re.sub(r"^```json\s*|\s*```$", "", content, flags=re.MULTILINE).strip()

        # ============================================================
        # ADDED: Extract token usage
        # ============================================================
        token_usage = {
            'prompt_tokens': response.usage.prompt_tokens,
            'completion_tokens': response.usage.completion_tokens,
            'total_tokens': response.usage.total_tokens
        }

        try:
            result = json.loads(cleaned)
            
            # Validate and normalize expected keys
            if "all_skills" not in result:
                result["all_skills"] = []
            
            if "priority_skills" not in result:
                result["priority_skills"] = result.get("all_skills", [])[:5]
            
            if "nice_to_have_skills" not in result:
                result["nice_to_have_skills"] = []
            
            if "linkedin_optimized_skills" not in result:
                result["linkedin_optimized_skills"] = result.get("all_skills", [])[:10]
                
            if "skill_categories" not in result:
                result["skill_categories"] = {}
                
            if "role_category" not in result:
                result["role_category"] = "Unknown"
                
            if "experience_level" not in result:
                result["experience_level"] = "Unknown"
            
            if "required_experience_years" not in result:
                result["required_experience_years"] = {}
                
            if "key_responsibilities" not in result:
                result["key_responsibilities"] = []
                
            if "qualifications" not in result:
                result["qualifications"] = []
            
            if "preferred_location" not in result:
                result["preferred_location"] = None
            
            # ============================================================
            # ADDED: Include token usage and model in result
            # ============================================================
            result['token_usage'] = token_usage
            result['model'] = 'gpt-4o-mini'
            
            print(f"✅ Skill extraction complete - Used {token_usage['total_tokens']} tokens")
            
            return result

        except json.JSONDecodeError as je:
            print(f"⚠️ JSON Decode Error: {je}")
            print(f"⚠️ Raw LLM Output: {content}")
            return get_default_error_response()

    except Exception as e:
        print(f"❌ Error calling OpenAI API: {e}")
        return get_default_error_response()

def get_default_error_response():
    '''Return default response when API fails'''
    return {
        "all_skills": ["Error extracting skills - please try again"],
        "priority_skills": [],
        "nice_to_have_skills": [],
        "skill_categories": {},
        "role_category": "Unknown",
        "experience_level": "Unknown",
        "required_experience_years": {},
        "key_responsibilities": [],
        "qualifications": [],
        "preferred_location": None
    }